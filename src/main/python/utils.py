from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font
import datetime, time
import pandas as pd
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment


# Function for flattening json
# https: // www.geeksforgeeks.org / flattening - json - objects - in -python /


def flatten_json(y):
    out = {}

    def flatten(x, name=""):

        # If the Nested key-value
        # pair is of dict type
        if type(x) is dict:

            for a in x:
                flatten(x[a], name + a + "_")

                # If the Nested key-value
        # pair is of list type
        elif type(x) is list:

            i = 0

            for a in x:
                flatten(a, name + str(i) + "_")
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


def fit_width(ws):
    dim_holder = DimensionHolder(worksheet=ws)
    dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=40)
    for col in range(ws.min_column + 1, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws, min=col, max=col, collapsed=True, bestFit=True
        )
    ws.column_dimensions = dim_holder


def write_data_frame_to_rows(ws, data, header=True):
    for r in dataframe_to_rows(data, index=False, header=header):
        ws.append(r)
    fit_width(ws)


def alignment_wrapper(ws):
    for row in ws.iter_rows():
        cell = row[:1][0]
        cell.alignment = Alignment(wrap_text=True, vertical='top')


def format_ws(ws):
    header = DifferentialStyle(font=Font(bold=True))
    r0 = Rule(type="expression", dxf=header)
    last_row = ws.max_row
    last_column = get_column_letter(ws.max_column)
    ws.conditional_formatting.add(f"A1:{last_column}1", r0)
    create_table(ws, ws.title, f"A1:{last_column}{last_row}")
    alignment_wrapper(ws)


def replace_illegal_characters(df):
    return df.applymap(
        lambda x: re.sub(ILLEGAL_CHARACTERS_RE, "", x) if isinstance(x, str) else x
    )


def create_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    # I list out the 4 show-xyz options here for reference
    style = TableStyleInfo(
        name="TableStyleMedium3",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def create_unix_timestamps(days=365):
    today = datetime.date.today()
    unixtime_today = time.mktime(today.timetuple())
    years_before = today - datetime.timedelta(days=days)
    unix_time_before = time.mktime(years_before.timetuple())
    return int(unixtime_today), int(unix_time_before)


def create_time_period_in_ymd_format(days=365):
    today = datetime.date.today().strftime("%Y-%m-%d")
    year_before = datetime.date.today() - datetime.timedelta(days)
    year_before = year_before.strftime("%Y-%m-%d")
    return today, year_before


def validate_data_frame(data: pd.DataFrame):
    if isinstance(data, pd.DataFrame) and not data.empty:
        return data
    else:
        raise ValueError("Data Frame doesn't exist or it's emtpy")


def upper_headers(data):
    try:
        data.columns = [str(col).title() for col in data.columns]
        return data
    except ValueError:
        print("Bad column value type")
